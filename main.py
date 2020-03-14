# 导入常用操作模块
import common
# 导入日期时间处理模块
import datetime
import time
# 导入正则表达式模块
import re

# 导入地理编码模块
import compare_location as CL
# 导入OpenPyXL处理EXCEL的xlsx文件
import openpyxl

# 处理考勤打卡记录
##################################################################################
# 数据源文件：IN考勤签卡数据.xlsx
# 字段：1.部门，2.员工编号，3.姓名，4.签卡时间，5.数据来源，6.创建时间

# 构建考勤签卡列表OCR_list (official checking recode)
# 字段：id, check_time, source
OCR_list = []

# 读入考勤签卡记录关键字段至OCR_list中
wb_OCR = openpyxl.load_workbook(filename="data/IN考勤签卡数据.xlsx")
ws_OCR = wb_OCR.active
row_index = 1
# 从第2行开始扫描记录
for per_row in ws_OCR.iter_rows():
    if row_index > 1:
        department = per_row[0].value
        id = per_row[1].value
        name = per_row[2].value
        # 若为光头数据，默认为考勤机机卡，补充数据
        if len(per_row[3].value) == 16:
            per_row[3].value += "(考勤机)"
        # 根据签卡类型，处理check_time
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "定点签卡":
            check_time = time.strptime(per_row[3].value[:16], "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "考勤机":
            check_time = time.strptime(per_row[3].value[:16], "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "外勤签卡":
            check_time = time.strptime(per_row[3].value[:16], "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "上班卡":
            check_time = time.strptime(per_row[3].value[:10] + " 08:59", "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "下班卡":
            check_time = time.strptime(per_row[3].value[:10] + " 17:31", "%Y-%m-%d %H:%M")

        source = re.findall(r'[(](.*?)[)]', per_row[4].value)[0]
        OCR_list.append([department, id, name, check_time, source])

    row_index += 1

# print("考勤签卡记录")
# for i in OCR_list: print(i)


# 生成考勤记录列表
##################################################################################
check_list = []
# 处理考勤签卡数据，将同一人、同一天的数据聚合成一条记录
# 数据源文件：IN考勤签卡数据.xlsx
# 字段：1.部门，2.员工编号，3.姓名，4.签卡时间，5.数据来源，6.创建时间
# 关键字段:1.部门 2.ID 3.姓名 4.类型（考勤/外勤） 5.签卡(check_time)[列表]
for i in range(0, len(OCR_list)):
    department = OCR_list[i][0]
    id = OCR_list[i][1]
    name = OCR_list[i][2]
    date = "%4d-%02d-%02d" % (OCR_list[i][3].tm_year, OCR_list[i][3].tm_mon, OCR_list[i][3].tm_mday)
    # 若列表为空，直接添加记录
    if len(check_list) == 0:
        check_list.append([department, id, name, date, "考勤", []])
        continue
    # 先扫描一遍签卡记录看看有没有同一人、同一天
    found = False
    for j in range(0, len(check_list)):
        if check_list[j][1] == id and check_list[j][3] == date:
            found = True
    if found == False:
        check_list.append([department, id, name, date, "考勤", []])
# 此时，check_list中已是聚合后的列表

# print("同人同日期聚合后的check_list")
# for i in check_list: print(i)


# 再扫描打卡记录，将时间写入聚合后的列表中
for i in range(0, len(OCR_list)):
    for j in range(0, len(check_list)):
        date = "%4d-%02d-%02d" % (OCR_list[i][3].tm_year, OCR_list[i][3].tm_mon, OCR_list[i][3].tm_mday)
        if check_list[j][1] == OCR_list[i][1] and check_list[j][3] == date:
            check_list[j][5].append("%02d:%02d" % (OCR_list[i][3].tm_hour, OCR_list[i][3].tm_min))

# print("check_list")
# for i in check_list:
#     while i[1] == "1053":
#         print(i)
#         break;


# 处理外出记录单 和 外勤打卡记录
##################################################################################
# 依据外出记录单.xlsx比对IN外勤打卡记录.xlsx
# 比对项目包括：同一人、同一天、同一地点、2条签卡记录
# 若缺少任何一要素，即判断外勤记录失效，若均符合向check_list添加1条外勤记录
# department, id, name, date, "外勤", [到达时间,离开时间]

# 读入外勤打卡记录
# 数据源文件：IN外勤打卡记录.xlsx
# 字段：1.所在部门，2.员工编号，3.员工姓名，4.地点，5.签卡时间，6.备注，7.审核人，8.审核时间，9.状态，10.签卡设备
# 关键字段：1.员工编号 2.签卡时间 3.地点
# 构建外勤列表MCR_list (mobile checking recode)
# 字段：department, id, name, check_time, location
MCR_list = []

# 读入外勤打卡记录关键字段至MCR_list中
wb_MCR = openpyxl.load_workbook(filename="data/IN外勤打卡记录.xlsx")
ws_MCR = wb_MCR.active
row_index = 1
for per_row in ws_MCR.iter_rows():
    if row_index > 1:
        department = per_row[0].value
        id = per_row[1].value
        name = per_row[2].value
        check_time = time.strptime(per_row[4].value, "%Y-%m-%d %H:%M")
        location = per_row[3].value
        MCR_list.append([department, id, name, check_time, location])
    row_index += 1

# print("外勤打卡记录")
# for i in MCR_list: print(i)

# 读入外出记录单到OW_list中
# 数据源文件：IN外出记录单.xlsx
# 字段：
# 1.编号，2.姓名，3.部门，4.外出时间，5.外出类型，(需加入员工编号）
# 6.人员类别，7.外出单位，8.外出地址9.相关项目名称，10.相关项目编号，
# 11.拜访人，12.相关销售人员，13.相关合同编号，14.联系人，15.联系方式，
# 16.外出具体目的，17.外出结果，18.审批，19.创建时间，20.流程状态
OW_list = []
wb_outwork = openpyxl.load_workbook(filename="data/IN外出记录单.xlsx")
ws_outwork = wb_outwork.active

row_index = 1
for per_row in ws_outwork.iter_rows():
    if row_index > 3:
        department = per_row[3].value
        id = per_row[1].value
        name = per_row[2].value
        date = per_row[4].value
        location = per_row[8].value
        OW_list.append([department, id, name, date, location, []])
    row_index += 1
# print("外出记录单")
# for i in OW_list:
#     while i[1] == "52":
#         print(i)
#         break


# 比对外勤打卡记录MCR_list
for i in range(0, len(OW_list)):
    # print(OW_list[i])
    for j in range(0, len(MCR_list)):
        id = MCR_list[j][1]
        date = "%4d-%02d-%02d" % (MCR_list[j][3].tm_year, MCR_list[j][3].tm_mon, MCR_list[j][3].tm_mday)

        # if MCR_list[j][1] == OW_list[i][1] and date == OW_list[i][3] and CL.compare_location(MCR_list[j][4], OW_list[i][4], 500):
        #     print(MCR_list[j])
        if MCR_list[j][1] == OW_list[i][1] and date == OW_list[i][3] and CL.compare_location(MCR_list[j][4],
                                                                                             OW_list[i][4], 400):
            OW_list[i][5].append("%02d:%02d" % (MCR_list[j][3].tm_hour, MCR_list[j][3].tm_min))

# print("对比后的OW_list")
# for i in OW_list:
#     while i[1] == "52":
#         print(i)
#         break

# 添加OW_list中有效数据至check_list中
for i in range(0, len(OW_list)):
    # check_list : [department, id, name, date, "考勤", []
    # OW_list : [department, id, name, date, location, []]
    if len(OW_list[i][5]) > 1:
        check_list.append([OW_list[i][0], OW_list[i][1], OW_list[i][2], OW_list[i][3], "外勤", OW_list[i][5]])

# for i in check_list: print(i)
#
#
#
# 聚合check_list形成 day_list
##################################################################################
# day_list字段如下：
# department, id, name, date, rec, status, reason
# rec为记录签卡情况的列表，status(正常|异常），reason为异常原因（迟到｜早退｜缺勤）
# check_list : [department, id, name, date, "考勤", [] ]
day_list = []

for i in range(0, len(check_list)):

    department = check_list[i][0]
    id = check_list[i][1]
    name = check_list[i][2]
    date = check_list[i][3]

    # 如果day_list为空，则直接添加记录
    if len(day_list) < 1:
        day_list.append([department, id, name, date, [], "", ""])

    # 先扫描一遍day_list看看有没有同一人、同一天
    found = False
    for j in range(0, len(day_list)):
        if day_list[j][1] == id and day_list[j][3] == date:
            found = True
    if found == False:
        day_list.append([department, id, name, date, [], "", ""])
    # 至此已形成day_list的聚合表


    # 对check_list打卡时间列表进行预处理
    check_list[i][5].sort()
    n = len(check_list[i][5])
    if check_list[i][4] == "考勤":
        if n == 1 and int(check_list[i][5][0][0:2]) < 12:
            rec_item = "考勤(%s-XX:XX)" % (check_list[i][5][0])
        if n == 1 and int(check_list[i][5][0][0:2]) >= 12:
            rec_item = "考勤（XX:XX-%s)" % (check_list[i][5][0])
        if n > 1:
            rec_item = "考勤(%s-%s)" % (check_list[i][5][0], check_list[i][5][n - 1])
    if check_list[i][4] == "外勤":
        rec_item = "外勤(%s-%s)" % (check_list[i][5][0], check_list[i][5][n - 1])
    # print(check_list[i][2], check_list[i][3], rec_item)

    # 再次扫描day_list，如果同一人、同一天那么合并rec字段
    for j in range(0, len(day_list)):
        if day_list[j][1] == id and day_list[j][3] == date:
            day_list[j][4].append(rec_item)

# for i in day_list:
#     while i[1] == "1053":
#         print(i)
#         break

#
# 扫描获取考勤开始日期和结束日期
for i in range(0, len(day_list)):
    if i == 0:
        start_date = day_list[i][3]
        end_date = day_list[i][3]
    else:
        if day_list[i][3] < start_date:
            start_date = day_list[i][3]
        if day_list[i][3] > end_date:
            end_date = day_list[i][3]
# print(start_date, end_date)
# 根据开始和结束日期，生成日期范围列表date_list
date_list = common.get_dates_bytimes(start_date, end_date)

# 扫描获取user_list
# user_list : department, id, name
user_list = []
for i in range(0, len(day_list)):
    if len(user_list) == 0:
        user_list.append([day_list[i][0], day_list[i][1], day_list[i][2]])
    found = False
    for j in range(0, len(user_list)):
        if user_list[j][1] == day_list[i][1]:
            found = True
    if found == False:
        user_list.append([day_list[i][0], day_list[i][1], day_list[i][2]])

# # for i in date_list: print(i)
# # for i in user_list: print(i)

#
# 判断day_list中数据是否存在异常，此部分判断逻辑是考核的关键
for per_row in day_list:
    AM = "出勤"
    PM = "出勤"
    for per_item in per_row[4]:
        type = per_item[0:2]
        start_time = per_item[3:8]
        end_time = per_item[9:14]
        # print(start_time,end_time)
        if type == "考勤":
            if start_time > "09:05":
                AM = "<迟到>"
            if end_time < "17:30":
                PM = "<早退>"
            if start_time  == "XX:XX":
                AM = "<缺勤>"
            if end_time == "XX:XX":
                PM = "<缺勤>"
        if type == "外勤":
            if start_time < "13:00":
                AM = "出勤"
            if end_time > "14:00":
                PM = "出勤"
    if AM != "出勤" or PM != "出勤":
        if AM != "出勤":
            per_row[5] = "异常"
            per_row[6] += AM
        if PM != "出勤":
            per_row[5] = "异常"
            per_row[6] += PM


# 读请假文件，请假是以小时进行的，按此更新day_list
# 读入请假记录关键字段至AFL_list中（ask for leave)
AFL_list = []

wb_AFL = openpyxl.load_workbook(filename="data/IN请假单.xlsx")
ws_AFL = wb_AFL.active

row_index = 1
for per_row in ws_AFL.iter_rows():
    if row_index > 3:
        department = per_row[3].value
        id = per_row[1].value
        name = per_row[2].value
        type = per_row[6].value
        hst_time = time.strptime(per_row[7].value, "%Y-%m-%d %H:%M")
        het_time = time.strptime(per_row[8].value, "%Y-%m-%d %H:%M")
        AFL_list.append([department, id, name, type, hst_time,het_time])
    row_index += 1

# for i in AFL_list:
#     print(i)

# day_list字段如下：
# department, id, name, date, rec, status, reason
# rec为记录签卡情况的列表，status(正常|异常），reason为异常原因（迟到｜早退｜缺勤）
# for i in day_list:
#     print(i[4][0])
#     print(re.findall(r'\d\d:\d\d', i[4][0]))

for AFL_per_row in AFL_list:
    # 扫描day_list, 为每条day_list记录判断是否请假的依据
    hst_time = AFL_per_row[4]
    het_time = AFL_per_row[5]

    for row_index in range(0,len(day_list)):
        rec_str = ""
        M_flag = False
        # 处理外勤记录特例，获取rec_str
        for per_item in day_list[row_index][4]:
            # print(day_list[row_index])
            # print(per_item)
            if len(per_item) < 15:
                break
            type = per_item[0:2]
            start_time = per_item[3:8]
            end_time = per_item[9:14]

            if type == "外勤":
                if start_time < "13:00":
                    rec_str += "08:59"
                if end_time > "14:00":
                    rec_str += "17:31"
            rec_str += per_item
        # 获取rec_str
        # print(rec_str)
        rec = re.findall(r'\d\d:\d\d', rec_str)
        rec.sort()
        st = rec[0]
        et = rec[len(rec)-1]

        # 转换成TIME结构
        rst_time = time.strptime(day_list[row_index][3] + " " + st, "%Y-%m-%d %H:%M")
        ret_time = time.strptime(day_list[row_index][3] + " " + et, "%Y-%m-%d %H:%M")
        if rst_time < hst_time:
            st_time = rst_time
        else:
            st_time = hst_time
        if ret_time > het_time:
            et_time = ret_time
        else:
            et_time = het_time
        # 输出st_time,et_time
        # print(day_list[row_index][3],time.strftime("%Y-%m-%d %H:%M",st_time),time.strftime("%Y-%m-%d %H:%M",et_time))

        # 比较最小时间st_time和最大时间et_time与当天的考勤时间是否符合
        if st_time < time.strptime(day_list[row_index][3] + " 09:05", "%Y-%m-%d %H:%M") and et_time >= time.strptime(day_list[row_index][3] + " 17:30", "%Y-%m-%d %H:%M"):
            M_flag = True
            # for i in day_list[row_index][4]:
            #     if i == "年假" or i == "事假" or i == "倒休":
            #         found = True

        if M_flag == True:
            # print(AFL_per_row[3])
            #print(day_list[row_index][4])
            day_list[row_index][4].append(AFL_per_row[3])
            day_list[row_index][5] = ""











# for i in day_list:
#     print(i)

# 在此处导出异常记录到文件中
# for i in day_list:
#     while i[5] == "异常":
#         print(i)
#         break




# for i in day_list:
#     while i[1] == "1053":
#         print(i)
#         break
#
####################################################################
# 输出已经过检验的day_list列表至EXCLE xlsx文件中
####################################################################

# 创建考勤报表工作簿
wb_report = openpyxl.Workbook()
ws_report = wb_report.active
# 设置工作表格式


# 写入表头
header = ["部门", "员工编号", "姓名"] + date_list
ws_report.append(header)

# 写入用户信息
for per_row in user_list:
    ws_report.append(per_row)

# 读day_list列表，将信息写入单元格
# 定位单元格行row 列col

for per_row in day_list:
    for row_index in range(1,ws_report.max_row+1):
        if per_row[1] == ws_report.cell(row_index,2).value:
            break
    for col_index in range(4,ws_report.max_column+1):
        if per_row[3] == ws_report.cell(1,col_index).value:
            break
    # print(row_index,col_index)

    # 拼接rec信息
    cell_str = ""
    for rec_item in per_row[4]:
        cell_str += rec_item
        cell_str += "\n"
    # 判断记录异常情况
    if per_row[5] == "异常":
        cell_str = per_row[6] + "\n" + cell_str
    # 写入单元格信息
    ws_report.cell(row_index,col_index).value = cell_str



# 标注单元格颜色
orange_fill = openpyxl.styles.PatternFill(fgColor="FFA500",fill_type='solid' )
yellow_fill = openpyxl.styles.PatternFill(fgColor="FFFF00",fill_type='solid' )
blue_fill = openpyxl.styles.PatternFill( fgColor="6495ED",fill_type='solid')


# 扫描ws_report表，标注休息日
for col_index in range(4,ws_report.max_column+1):
    date_header = time.strptime(ws_report.cell(1,col_index).value, "%Y-%m-%d")
    # tm_wday 取值0-6，0是周一，1是周二，2是周三，3是周四，4是周五，5是周六，6是周日
    # 如果是周末，那就标注为休息并将单元格标为蓝色
    if date_header.tm_wday ==5 or date_header.tm_wday == 6:
        for row_index in range(2,ws_report.max_row+1):
            if ws_report.cell(row_index,col_index).value ==  None:
                ws_report.cell(row_index,col_index).value = "休息"
            else:
                ws_report.cell(row_index,col_index).value = "休息\n" + str(ws_report.cell(row_index,col_index).value)
            ws_report.cell(row_index, col_index).fill = blue_fill

# 扫描ws_report表，如果单元格内容为空，则意味着缺少考勤记录标记为橙色
for row_index in range(1,ws_report.max_row+1):
    for col_index in range(4,ws_report.max_column+1):
        cell_str = str(ws_report.cell(row_index,col_index).value)
        # print(row_index,col_index,"-",cell_str)
        if cell_str == "None":
            ws_report.cell(row_index,col_index).value = "缺勤"
            ws_report.cell(row_index,col_index).fill = yellow_fill
        if cell_str.find("<") != -1:
            ws_report.cell(row_index, col_index).fill = orange_fill
        # if cell_str.find("XX:XX") != -1:
        #     ws_report.cell(row_index, col_index).fill = orange_fill
        # if cell_str.find("<早退>") != -1:
        #     ws_report.cell(row_index, col_index).fill = orange_fill
        # if cell_str.find("<迟到>") != -1:
        #     ws_report.cell(row_index, col_index).fill = orange_fill












# 向考勤报表工作簿中写入考勤信息
output_file = "data/OUT考勤报表.xlsx"
wb_report.save(filename=output_file)


