from datetime import date, timedelta, datetime


def days_cur_month():
    m = datetime.now().month
    y = datetime.now().year
    ndays = (date(y, m + 1, 1) - date(y, m, 1)).days
    d1 = date(y, m, 1)
    d2 = date(y, m, ndays)
    delta = d2 - d1

    return [(d1 + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]


def trans_checking_in(cell_str):
# 输入情形：
# 正常班：2次打卡时间
# 有外勤：0次 / 1次打卡时间 / 外出 （此种情况不会在报表中出现）
# 异常情况：年假 / 事假 / 病假 / 出差 / 调休

# 输出格式：
#     常班XX: XX - XX:XX
#     外勤XX: XX - XX:XX
#     年假XX: XX - XX:XX
#     事假XX: XX - XX:XX
#     病假XX: XX - XX:XX
#     出差XX: XX - XX:XX
#     调休XX: XX - XX:XX
#     若全天，则时间标记为09:00-17:30
#     no_space_str = str(cell_str).replace(' ','')
    if str(cell_str).find("年假") != -1:
        return "年假09:00-17:30"
    if str(cell_str).find("事假") != -1:
        return "事假09:00-17:30"
    if str(cell_str).find("病假") != -1:
        return "病假09:00-17:30"
    if str(cell_str).find("出差") != -1:
        return "出差09:00-17:30"
    if str(cell_str).find("调休") != -1:
        return "调休09:00-17:30"
# modify
    return str(cell_str)


# 导入OpenPyXL处理EXCEL的xlsx文件
from openpyxl import Workbook
from openpyxl import load_workbook

# 设置相关文件名称
base_info_file = "data/base_info.xlsx"
dest_file = "data/考勤月报.xlsx"

# 创建考勤月报工作簿
wb_dest = Workbook()
ws_dest = wb_dest.active
# 写入表头
header_dest = ["部门", "员工编号", "姓名"] + days_cur_month()

ws_dest.append(header_dest)

# 读入基础信息工作簿，构建people列表
wb_base_info = load_workbook(filename=base_info_file)
ws_base_info = wb_base_info.active
# rows_base_info = ws_base_info.max_row

row_index = 1
people = []
for per_row in ws_base_info.iter_rows():
    if row_index > 1:
        people.append([per_row[0].value, per_row[1].value, per_row[2].value])
    row_index += 1
print(people)

# 向考勤月报工作簿中写入基础信息

for per_people in people:
    print(per_people)
    ws_dest.append(per_people)

#######################################################
# 填充数据，核心代码
# 1.读入常驻地usual place of residence (UPOR)考勤工作簿，扫描所有单元格，并构建写入内容
UPOR_file = "data/常驻地考勤.xlsx"
wb_UPOR = load_workbook(filename=UPOR_file)
ws_UPOR = wb_UPOR.active


for row in range(2, ws_UPOR.max_row+1):
    for col in range(4, ws_UPOR.max_column+1):
        ws_dest.cell(row, col).value = trans_checking_in(ws_UPOR.cell(row, col).value)

#
#         # ws_dest.cell(row,col).value = str(row)+'-'+str(col)
# a = list(range(2, ws_UPOR.max_row+1))
# b = list(range(4, ws_UPOR.max_column+1))
# print(a)
# print(b)
######################################################


# 保存EXCEL文件
wb_UPOR.save(filename=UPOR_file)
wb_dest.save(filename=dest_file)
